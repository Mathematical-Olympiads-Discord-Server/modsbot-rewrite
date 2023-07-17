import ast
from functools import reduce
import math
import openpyxl.utils
import random
import statistics
from datetime import datetime, timedelta
import re
from typing import Optional

import discord
import schedule
import threading
import asyncio
import io
import aiohttp

from discord import app_commands
from discord.ext import commands
from discord.ext.commands import BucketType

from cogs import config as cfg

Cog = commands.Cog

POTD_RANGE = 'POTD!A2:S'
CURATOR_RANGE = 'Curators!A3:E'

days = [None, 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']


def is_pc(ctx):
    if ctx.guild is None:
        return False
    return cfg.Config.config['problem_curator_role'] in [x.id for x in ctx.author.roles]

async def dm_or_channel(user: discord.User, channel: discord.abc.Messageable, content='', *args, **kargs):
    try:
        if user is not None and not user.bot:
            await user.send(*args, content=content, **kargs)
    except Exception:
        await channel.send(*args, content=user.mention+'\n'+content, **kargs)


class Potd(Cog):

    def __init__(self, bot: commands.Bot):

        self.listening_in_channel = -1
        self.to_send = ''
        self.bot = bot
        self.ping_daily = False
        self.late = False
        self.requested_number = -1
        self.dm_list = []
        self.timer = None

        reply = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=POTD_RANGE).execute()
        values = reply.get('values', [])
        self.latest_potd = int(values[0][0])  
        
        cursor = cfg.db.cursor()
        cursor.execute('''INSERT OR IGNORE INTO settings (setting, value) VALUES
            ('potd_dm', 'True')
            ''')
        cfg.db.commit()
        cursor.execute("SELECT value FROM settings WHERE setting = 'potd_dm'")
        self.enable_dm = (cursor.fetchone()[0] == 'True')

        schedule.every().day.at("10:00").do(self.schedule_potd).tag('cogs.potd')
        schedule.every().day.at("09:00").do(lambda: self.schedule_potd(1)).tag('cogs.potd')
        schedule.every().day.at("07:00").do(lambda: self.schedule_potd(3)).tag('cogs.potd')
        schedule.every().day.at("04:00").do(lambda: self.schedule_potd(6)).tag('cogs.potd')
        schedule.every().day.at("22:00").do(lambda: self.schedule_potd(12)).tag('cogs.potd')

        schedule.every().hour.at("10:00").do(self.post_proposed_potd).tag('cogs.potd.proposal')


    @commands.command()
    @commands.check(is_pc)
    async def reset_potd(self, ctx=None):
        self.requested_number = -1
        self.listening_in_channel = -1
        self.to_send = ''
        self.late = False
        self.ping_daily = False
        self.dm_list = []
        try:
            self.timer.cancel()
        except Exception:
            pass
        self.timer = None

    def reset_if_necessary(self):
        if self.listening_in_channel != -1:
            self.bot.loop.create_task(self.reset_potd())

    def prepare_dms(self, potd_row):
        def should_dm(x):
            for i in range(4):
                if (['a', 'c', 'g', 'n'][i] in potd_row[5].lower()) and not (x[1][4*i] == 'x'):
                    if int(x[1][4*i:4*i+2]) <= d <= int(x[1][4*i+2:4*i+4]):
                        return True
            return False

        try:
            d = int(potd_row[6])
        except Exception:
            return

        cursor = cfg.db.cursor()
        cursor.execute("SELECT * FROM potd_ping2")
        result = cursor.fetchall()
        self.dm_list = [i[0] for i in filter(should_dm, result)]

    def curator_id(self, curators, value):
        value = str(value)
        if value == '':
            return None
        for i in curators:
            for j in range(min(len(i), 4)):
                if value == str(i[j]):
                    return i[0]
        return None

    def generate_source(self, potd_row):
        # Figure out whose potd it is
        curators = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=CURATOR_RANGE).execute().get('values', [])
        curator_id = self.curator_id(curators, potd_row[3])
        if curator_id is None:
            curator = 'Unknown Curator'
        else:
            curator = f'<@!{curator_id}>'
        difficulty_length = len(potd_row[5]) + len(potd_row[6])
        padding = (' ' * (max(35 - len(potd_row[4]), 1)))

        source = discord.Embed()
        source.add_field(name='Curator', value=curator)
        source.add_field(name='Source', value=f'||`{potd_row[4]}{padding}`||')
        source.add_field(name='Difficulty', value=f'||`{str(potd_row[6]).ljust(5)}`||')
        source.add_field(name='Genre', value=f'||`{str(potd_row[5]).ljust(5)}`||')

        # Community Rating footer
        cursor = cfg.db.cursor()
        cursor.execute(f'SELECT * FROM ratings WHERE prob = {potd_row[0]}')
        result = cursor.fetchall()

        community_rating = ''
        if len(result) > 0:
            community_rating += f"There are {len(result)} community difficulty ratings. "
            try:
                underrate_count = sum(row[3] < int(potd_row[6]) for row in result)
                if underrate_count > 0:
                    community_rating += f"{underrate_count} rated lower than current rating. "
                overrate_count = sum(row[3] > int(potd_row[6]) for row in result)
                if overrate_count > 0:
                    community_rating += f"{overrate_count} rated higher than current rating. "            
            except:
                pass
            community_rating += "\n"
        
        # Final footer
        source.set_footer(text=f'{community_rating}Use -rating {potd_row[0]} to check the community difficulty rating of this problem '
                            f'or -rate {potd_row[0]} rating to rate it yourself. React with a 👍 if you liked '
                            f'the problem. ')

        return source
    
    async def edit_source(self, potd):
        sheet = self.get_potd_sheet()
        potd_row = self.get_potd_row(potd, sheet)
        try:
            potd_source = self.generate_source(potd_row)
            potd_source_msg_id = potd_row[cfg.Config.config['potd_sheet_message_id_col']]
            potd_source_msg = await self.bot.get_channel(cfg.Config.config['potd_channel']).fetch_message(potd_source_msg_id)
            await potd_source_msg.edit(embed=potd_source)
        except:
            pass

    def schedule_potd(self, mode=None):
        self.bot.loop.create_task(self.check_potd(mode))

    def responsible(self, potd_id:int, urgent:bool=False):     # Mentions of responsible curators

        # Get stuff from the sheet (API call)
        potds = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=POTD_RANGE).execute().get('values', [])
        curators = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=CURATOR_RANGE).execute().get('values', [])
        try:
            i = int(potds[0][0]) - int(potd_id)
        except Exception:
            return 'Invalid entry (A2) in spreadsheet! '
        potd_row = potds[i]

        # Searches for relevant curators
        mentions = ''
        r_list = []
        try:
            day = str(days.index(str(potd_row[2])))
        except Exception:
            return 'Day not recognized. '
        for curator in curators:
            try:
                if (curator[4] == day):
                    mentions += f'<@{curator[0]}> '
                    r_list.append(curator)
            except Exception:
                pass
        if urgent:
            return mentions + f'<@&{cfg.Config.config["problem_curator_role"]}> '
        if mentions == '':
            return f'No responsible curators found for the potd on {potd_row[1]}!'

        # Searches for curator whose last curation on this day of the week was longest ago.
        i += 7
        while (i < len(potds)) and (len(r_list) > 1):
            try:
                for curator in r_list:
                    if curator[0] == self.curator_id(curators, potds[i][3]):
                        r_list.remove(curator)
            except Exception:
                pass
            i += 7
        return f'<@{r_list[0][0]}> '

    async def potd_embedded(self, ctx, *, number: int):
        # It can only handle one at a time!
        if self.listening_in_channel != -1:
            await ctx.send("Please wait until the previous potd call has finished!")
            return

        reply = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=POTD_RANGE).execute()
        values = reply.get('values', [])
        current_potd = int(values[0][0])  # this will be the top left cell which indicates the latest added potd
        potd_row = values[current_potd - number]  # this gets the row requested

        # Create the message to send
        to_tex = ''
        try:
            to_tex = '<@419356082981568522>\n```tex\n \\textbf{Day ' + str(number) + '} --- ' + str(potd_row[2]) + ' ' + str(
                potd_row[1]) + '\\vspace{11pt}\\\\\\setlength\\parindent{1.5em}' + str(potd_row[8]) + '```'
        except IndexError:
            await ctx.send("There is no potd for day {}. ".format(number))
            return
        print(to_tex)

        # Finish up
        self.requested_number = int(potd_row[0])
        self.latest_potd = int(potd_row[0])
        self.to_send = self.generate_source(potd_row)
        self.listening_in_channel = ctx.channel.id
        self.late = True
        await ctx.send(to_tex, delete_after=20)

    async def check_potd(self, mode=None):

        # Get the potds from the sheet (API call)
        potds = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=POTD_RANGE).execute().get('values', [])

        # Check today's potd
        if mode is None:
            next = datetime.now()
            date = next.strftime("%d %b %Y")
            soon = [(next + timedelta(days = i)).strftime("%d %b %Y") for i in range(1, 4)]
        else:
            next = datetime.now() + timedelta(hours = mode)
            date = next.strftime("%d %b %Y")
            soon = [date]
        if date[0] == '0':
            date = date[1:]
        for i in range(len(soon)):
            if soon[i][0] == '0':
                soon[i] = soon[i][1:]
        passed_current = False
        potd_row = None
        fail = False
        remind = []
        curator_role = (await self.bot.fetch_guild(cfg.Config.config['mods_guild'])).get_role(cfg.Config.config['problem_curator_role'])
        j = 1                   # TESTING
        for potd in potds:
            j += 1              # TESTING
            if len(potd) < 2:   # TESTING
                await self.bot.get_channel(cfg.Config.config['log_channel']).send(
                        f"Invalid entry at row {j}, potd = {potd}")
                pass
            if passed_current:
                if len(potd) < 8:  # Then there has not been a potd on that day.
                    fail = True
                    await curator_role.edit(mentionable = True)
                    await self.bot.get_channel(cfg.Config.config['helper_lounge']).send(
                        f"There was no potd on {potd[1]}! {self.responsible(int(potd[0]), True)}")
                    await curator_role.edit(mentionable = False)
            if potd[1] == date:
                passed_current = True
                potd_row = potd
                if len(potd) < 8 and (mode is None):  # There is no potd.
                    fail = True
                    await curator_role.edit(mentionable = True)
                    await self.bot.get_channel(cfg.Config.config['helper_lounge']).send(
                        f"There is no potd today! {self.responsible(int(potd[0]), True)}")
                    await curator_role.edit(mentionable = False)
            if potd[1] in soon:
                if len(potd) < 8:  # Then there is no potd on that day.
                    remind.append(int(potd[0]))
                soon.remove(potd[1])
        if soon != []:
            await self.bot.get_channel(cfg.Config.config['helper_lounge']).send(
                f"Insufficient rows in the potd sheet! ")
        if remind != []:
            mentions = ''
            for i in remind:
                mentions += self.responsible(i, (mode == 1) or (mode == 3))
            await curator_role.edit(mentionable = True)
            await self.bot.get_channel(cfg.Config.config['helper_lounge']).send(
                f"Remember to fill in your POTDs! {mentions}")
            await curator_role.edit(mentionable = False)
        if fail or not (mode is None):
            return

        print('l123')
        # Otherwise, everything has passed and we are good to go.
        # Create the message to send
        to_tex = '<@419356082981568522>\n```tex\n\\textbf{Day ' + str(potd_row[0]) + '} --- ' + str(potd_row[2]) + ' ' + str(
            potd_row[1]) + '\\vspace{11pt}\\\\\\setlength\\parindent{1.5em}' + str(potd_row[8]) + '```'
        print(to_tex)

        # Finish up
        self.requested_number = int(potd_row[0])
        self.latest_potd = int(potd_row[0])
        self.prepare_dms(potd_row)
        self.to_send = self.generate_source(potd_row)
        self.listening_in_channel = cfg.Config.config['potd_channel']
        self.ping_daily = True
        self.late = False
        await self.bot.get_channel(cfg.Config.config['potd_channel']).send(to_tex, delete_after=20)
        await self.create_potd_forum_post(self.requested_number)
        print('l149')
        # In case Paradox unresponsive
        self.timer = threading.Timer(20, self.reset_if_necessary)
        self.timer.start()

    async def create_potd_forum_post(self, number):
        forum = self.bot.get_channel(cfg.Config.config['potd_forum'])
        await forum.create_thread(name=f"POTD {number}", content="potd")

    @Cog.listener()
    async def on_message(self, message: discord.Message):
        if message.channel.id == self.listening_in_channel and int(message.author.id) == cfg.Config.config[
            'paradox_id']:
            
            # m = await message.channel.send(
            #     '{} \nRate this problem with `-rate {} <rating>` and check its user difficulty rating with `-rating {}`'.format(
            #         self.to_send, self.requested_number, self.requested_number))
            self.listening_in_channel = -1 # Prevent reset
            source_msg = await message.channel.send(embed=self.to_send) 
            await source_msg.add_reaction("👍")
            if self.late:
                await source_msg.add_reaction('⏰')

            if message.channel.id == cfg.Config.config['potd_channel']:
                # record the ID of the source_msg if it is in POTD channel 
                # get the row and column to update
                column = openpyxl.utils.get_column_letter(cfg.Config.config['potd_sheet_message_id_col']+1)
                reply = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=POTD_RANGE).execute()
                values = reply.get('values', [])
                current_potd = int(values[0][0])  # this will be the top left cell which indicates the latest added potd
                row = current_potd - self.requested_number + 2  # this gets the row requested
                # update the source_msg in the sheet
                request = cfg.Config.service.spreadsheets().values().update(spreadsheetId=cfg.Config.config['potd_sheet'], 
                                                            range=f'{column}{row}', valueInputOption='RAW',body={"range": f'{column}{row}', "values": [[str(source_msg.id)]] })
                response = request.execute()

                # record the link to rendered image if it is in POTD channel 
                # get the row and column to update
                column = openpyxl.utils.get_column_letter(cfg.Config.config['potd_sheet_image_link_col']+1)
                reply = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=POTD_RANGE).execute()
                values = reply.get('values', [])
                current_potd = int(values[0][0])  # this will be the top left cell which indicates the latest added potd
                row = current_potd - self.requested_number + 2  # this gets the row requested
                # update the source_msg in the sheet
                request = cfg.Config.service.spreadsheets().values().update(spreadsheetId=cfg.Config.config['potd_sheet'], 
                                                            range=f'{column}{row}', valueInputOption='RAW',body={"range": f'{column}{row}', "values": [[str(message.attachments[0].proxy_url)]] })
                response = request.execute()

            bot_log = self.bot.get_channel(cfg.Config.config['log_channel'])

            ping_msg = None
            if self.ping_daily:
                r = self.bot.get_guild(cfg.Config.config['mods_guild']).get_role(cfg.Config.config['potd_role'])
                await r.edit(mentionable=True)
                ping_msg = await message.channel.send('<@&{}>'.format(cfg.Config.config['potd_role']))
                await r.edit(mentionable=False)

                if self.enable_dm:

                    bot_spam = self.bot.get_channel(cfg.Config.config['bot_spam_channel'])
                    potd_discussion_channel = self.bot.get_channel(cfg.Config.config['potd_discussion_channel'])

                    ping_embed = discord.Embed(title=f'POTD {self.latest_potd} has been posted: ',
                        description=f'{potd_discussion_channel.mention}\n{message.jump_url}', colour=0xDCDCDC)
                    for field in self.to_send.to_dict()['fields']:
                        ping_embed.add_field(name=field['name'], value=field['value'])
                    if message.attachments == []:
                        await bot_log.send('No attachments found! ')
                    else:
                        ping_embed.set_image(url=message.attachments[0].url)
                        dm_failed = []
                        for id in self.dm_list:
                            user = self.bot.get_user(int(id))
                            try:
                                await user.send(embed=ping_embed)
                            except Exception:
                                dm_failed.append(id)
                        if dm_failed != []:
                            msg = 'Remember to turn on DMs from this server to get private notifications! '
                            for id in dm_failed: msg += f'<@{id}> '
                            await bot_spam.send(msg, embed=ping_embed)

            if message.channel.id == cfg.Config.config['potd_channel']:
                try:
                    await message.publish()
                    await source_msg.publish()
                except Exception:
                    await bot_log.send('Failed to publish!')

            cursor = cfg.db.cursor()
            if ping_msg is None:
                cursor.execute(f'''INSERT INTO potd_info (potd_id, problem_msg_id, source_msg_id, ping_msg_id) VALUES
                    ('{self.latest_potd}', '{message.id}', '{source_msg.id}', '')''')
            else:
                cursor.execute(f'''INSERT INTO potd_info (potd_id, problem_msg_id, source_msg_id, ping_msg_id) VALUES
                    ('{self.latest_potd}', '{message.id}', '{source_msg.id}', '{ping_msg.id}')''')
            cfg.db.commit()

            await self.reset_potd()
            await bot_log.send('POTD execution successful.')

    @commands.command(aliases=['potd'], brief='Displays the potd with the provided number. ')
    @commands.check(is_pc)
    async def potd_display(self, ctx, number: int):

        # It can only handle one at a time!
        if self.listening_in_channel != -1:
            await dm_or_channel(ctx.author, self.bot.get_channel(cfg.Config.config['helper_lounge']),
                "Please wait until the previous call has finished!")
            return

        # Read from the spreadsheet
        reply = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=POTD_RANGE).execute()
        values = reply.get('values', [])
        current_potd = int(values[0][0])  # this will be the top left cell which indicates the latest added potd
        potd_row = values[current_potd - number]  # this gets the row requested

        # Create the message to send
        to_tex = ''
        try:
            to_tex = '<@419356082981568522>\n```tex\n\\textbf{Day ' + str(potd_row[0]) + '} --- ' + str(potd_row[2]) + ' ' + str(
                potd_row[1]) + '\\vspace{11pt}\\\\\\setlength\\parindent{1.5em}' + str(potd_row[8]) + '```'
        except IndexError:
            await dm_or_channel(ctx.author, self.bot.get_channel(cfg.Config.config['helper_lounge']),
                f"There is no potd for day {number}. ")
            return
        print(to_tex)

        # Finish up
        self.requested_number = int(potd_row[0])
        self.latest_potd = int(potd_row[0])
        self.prepare_dms(potd_row)
        self.to_send = self.generate_source(potd_row)
        self.listening_in_channel = ctx.channel.id
        self.late = True
        self.ping_daily = False
        await ctx.send(to_tex, delete_after=20)
        # In case Paradox unresponsive
        self.timer = threading.Timer(20, self.reset_if_necessary)
        self.timer.start()

    @commands.command(aliases=['fetch'], brief='Fetch a potd by id.',
                      help='`-fetch 1`: Fetch POTD Day 1.\n'
                            '`-fetch 1 s`: Fetch POTD Day 1, masked by spoiler.\n'
                            '`-fetch 1 t`: Fetch POTD Day 1, in tex form.\n')
    @commands.cooldown(1, 5, BucketType.user)
    async def potd_fetch(self, ctx, number: int, flag: str=''):
        sheet = self.get_potd_sheet()
        potd_row = self.get_potd_row(number, sheet)

        if potd_row == None:
            await ctx.send(f"There is no potd for day {number}. ")
            return
        else:
            # Create the message to send
            try:
                # if there is image link, just send it out
                image_link = self.check_for_image_link(potd_row)
                if image_link and 't' not in flag:
                    async with aiohttp.ClientSession() as session:
                        async with session.get(image_link) as resp:
                            if resp.status != 200:
                                return await ctx.send('Could not download file...')
                            data = io.BytesIO(await resp.read())
                            if 's' not in flag:
                                await ctx.send(file=discord.File(data, f'potd{number}.png'))
                            else:
                                await ctx.send(file=discord.File(data, f'SPOILER_potd{number}.png'))
                # if no image link, send tex
                else:
                    if 's' not in flag:
                        output = '<@' + str(cfg.Config.config['paradox_id']) + '>\n' + self.texify_potd(potd_row)
                    else:
                        output = '<@' + str(cfg.Config.config['paradox_id']) + '>texsp\n||' + self.texify_potd(potd_row) + '||'
                    await ctx.send(output, delete_after=5)
            except IndexError:
                await ctx.send(f"There is no potd for day {number}. ")
                return

    def check_for_image_link(self, potd_row) -> Optional[str]:
        if len(potd_row) >= 19 and potd_row[cfg.Config.config['potd_sheet_image_link_col']] not in [None, '']:
            return potd_row[cfg.Config.config['potd_sheet_image_link_col']]
        else:
            return None

    def texify_potd(self, potd_row) -> str:
        return '```tex\n\\textbf{Day ' + str(
            potd_row[cfg.Config.config['potd_sheet_id_col']]) + '} --- ' + str(
            potd_row[cfg.Config.config['potd_sheet_day_col']]) + ' ' + str(
            potd_row[cfg.Config.config['potd_sheet_date_col']]) + '\\vspace{11pt}\\\\\\setlength\\parindent{1.5em}' + str(
            potd_row[cfg.Config.config['potd_sheet_statement_col']]) + '```'

    @commands.command(aliases=['source'], brief='Get the source of a potd by id.')
    @commands.cooldown(1, 5, BucketType.user)
    async def potd_source(self, ctx, number: int):
        sheet = self.get_potd_sheet()
        potd_row = self.get_potd_row(number, sheet)

        if potd_row == None:
            await ctx.send(f"There is no potd for day {number}. ")
            return
        else:
            source = self.generate_source(potd_row)
            await ctx.send(embed=source)

    @commands.command(aliases=['search'], brief='Search for a POTD by genre and difficulty.',
        help='`-search 4 6`: Search for a POTD with difficulty d4 to d6 (inclusive).\n'
            '`-search 4 6 C`: Search for a POTD with difficulty d4 to d6 and genres including combinatorics.\n'
            '`-search 4 6 CG`: Search for a POTD with difficulty d4 to d6 and genres including combinatorics or geometry.\n'
            '`-search 4 6 \'CG\'`: Search for a POTD with difficulty d4 to d6 and genres including (combinatorics AND geometry).\n'
            '`-search 4 6 A\'CG\'N`: Search for a POTD with difficulty d4 to d6 and genres including (algebra OR (combinatorics AND geometry) OR number theory).\n'
            '`-search 4 6 ACGN false`: Search for a POTD with difficulty d4 to d6. Allow getting problems marked in the `-solved` list.')
    @commands.cooldown(1, 5, BucketType.user)
    async def potd_search(self, ctx, diff_lower_bound:int, diff_upper_bound:int, genre:str='ACGN', search_unsolved:bool=True):
        if diff_lower_bound > diff_upper_bound:
            await ctx.send(f"Difficulty lower bound cannot be higher than upper bound.")
            return

        # Set up the genre filter
        genre_filter = self.parse_genre_input(genre)

        # set up the difficulty filter
        diff_lower_bound_filter = max(0,diff_lower_bound)
        diff_upper_bound_filter = max(min(99, diff_upper_bound), diff_lower_bound_filter)
        
        potds = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=POTD_RANGE).execute().get('values', [])
        picked_potd = self.pick_potd(diff_lower_bound_filter, diff_upper_bound_filter, genre_filter, potds, [], ctx, search_unsolved)
        if picked_potd is not None:
            # fetch the picked POTD
            await self.potd_fetch(ctx, int(picked_potd))
        else:
            await ctx.send(f"No POTD found!")

    def potds_filtered_by_keywords(self, keyword_list: list[str]):
        potds = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=POTD_RANGE).execute().get('values', [])
        filtered_potds = [x for x in potds if len(x) > cfg.Config.config['potd_sheet_statement_col']
                        and all(keyword.lower() in x[cfg.Config.config['potd_sheet_statement_col']].lower() for keyword in keyword_list)]
        return filtered_potds

    async def potd_search_keywords_autocomplete(self, interaction: discord.Interaction, current: str) -> list[app_commands.Choice[str]]:
        filtered_potds = self.potds_filtered_by_keywords(current.split())
        filtered_potd_statements = [potd[cfg.Config.config['potd_sheet_statement_col']] for potd in  filtered_potds]
        # Only 25 responses are supported in autocomplete, and they must be at most 100 characters
        return [app_commands.Choice(name=statement[:100], value=statement[:100]) for statement in filtered_potd_statements][:25]
    
    @app_commands.command()
    @app_commands.describe(keywords='Search past potds using these keywords')
    @app_commands.autocomplete(keywords=potd_search_keywords_autocomplete)
    @commands.cooldown(1, 10, BucketType.user)
    async def potd_keywords(self, interaction: discord.Interaction, keywords: str):
        """Search potds using keywords"""

        filtered_potds = self.potds_filtered_by_keywords(keywords.split())
            
        if filtered_potds:
            picked_potd_row = random.choice(filtered_potds)
            image_link = self.check_for_image_link(picked_potd_row)
            if image_link:
                await interaction.response.send_message(f"[image]({image_link})")
            else:
                output = '<@' + str(cfg.Config.config['paradox_id']) + '>\n' + self.texify_potd(picked_potd_row)
                await interaction.response.send_message(output, delete_after=5)
        else:
            await interaction.response.send_message(f"No POTD found!", ephemeral=True)

    def parse_genre_input(self, genre):
        complex_genres = genre.split("'")[1::2]
        simple_genres = "".join(genre.split("'")[0::2])

        genre_filter = []
        for character in simple_genres:
            if character.upper() == "A":
                genre_filter.append("A")
            if character.upper() == "C":
                genre_filter.append("C")
            if character.upper() == "G":
                genre_filter.append("G")
            if character.upper() == "N":
                genre_filter.append("N")

        for item in complex_genres:
            parsed_complex_genre = set()
            for character in item:
                if character.upper() == "A":
                    parsed_complex_genre.add("A")
                if character.upper() == "C":
                    parsed_complex_genre.add("C")
                if character.upper() == "G":
                    parsed_complex_genre.add("G")
                if character.upper() == "N":
                    parsed_complex_genre.add("N")
            parsed_complex_genre = "".join(parsed_complex_genre)
            genre_filter.append(parsed_complex_genre)

        return set(genre_filter)

    @commands.command(aliases=['mock'], brief='Create a mock paper using past POTDs.',
        help='`-mock IMO`: create mock IMO paper\n'
            '\n'
            'See below for a list of available templates and respective difficulty ranges\n'
            '(e.g. [5,7],[7,9],[9,11],[5,7],[7,9],[9,11] means problem 1 is d5-7, problem 2 is d7-9, etc.) \n'
            '\n'
            'IMO (International Mathematical Olympiad):\n'
            '[5,7],[7,9],[9,11],[5,7],[7,9],[9,11]\n'
            'AMO (Australian Mathematical Olympiad):\n'
            '[2,3],[3,4],[4,5],[5,6],[2,3],[3,4],[4,5],[5,6]\n'
            'APMO (Asian Pacific Mathematics Olympiad):\n'
            '[4,5],[5,6],[6,7],[7,8],[8,10]\n'
            'BMO1 (British Mathematical Olympiad Round 1):\n'
            '[1,2],[1,2],[2,3],[2,3],[3,5],[3,6]\n'
            'BMO2 (British Mathematical Olympiad Round 2):\n'
            '[3,4],[4,5],[5,6],[6,7]\n'
            'IGO (Iranian Geometry Olympiad):\n'
            '[5,6],[6,7],[7,8],[8,9],[9,10]\n'
            'NZMO2 (New Zealand Mathematical Olympiad Round 2):\n'
            '[1,2],[2,3],[3,4],[4,5],[5,6]\n'
            'SMO2 (Singapore Mathematical Olympiad Open Round 2):\n'
            '[4,5],[5,6],[6,7],[7,8],[8,9]\n'
            'USAMO (United States of America Mathematical Olympiad):\n'
            '[5,7],[7,9],[9,11],[5,7],[7,9],[9,11]\n'
            'USAJMO (United States of America Junior Mathematical Olympiad):\n'
            '[3,5],[5,7],[7,8],[3,5],[5,7],[7,8]\n'
            'CHINA (Crushingly Hard Imbalanced Nightmarish Assessment):\n'
            '[7,8],[8,10],[10,12],[7,8],[8,10],[10,12]')
    @commands.cooldown(1, 30, BucketType.user)
    async def potd_mock(self, ctx, template:str="IMO", search_unsolved:bool=True):
        template = template.upper()
        template_list = ["IMO", "AMO", "APMO", "BMO1", "BMO2", "IGO", "NZMO2", "SMO2", "USAMO", "USAJMO", "CHINA"]
        if template not in template_list and template != "AFMO":
            await ctx.send(f"Template not found. Possible templates: {', '.join(template_list)}. Use `-help potd_mock` for more details.")
            return
        else:
            if template == "IMO":
                difficulty_bounds = [[5,7],[7,9],[9,11],[5,7],[7,9],[9,11]]
            elif template == "AMO":
                difficulty_bounds = [[2,3],[3,4],[4,5],[5,6],[2,3],[3,4],[4,5],[5,6]]
            elif template == "APMO":
                difficulty_bounds = [[4,5],[5,6],[6,7],[7,8],[8,10]]
            elif template == "BMO1":
                difficulty_bounds = [[1,2],[1,2],[2,3],[2,3],[3,5],[3,6]]
            elif template == "BMO2":
                difficulty_bounds = [[3,4],[4,5],[5,6],[6,7]]         
            elif template == "IGO":
                difficulty_bounds = [[5,6],[6,7],[7,8],[8,9],[9,10]]
            elif template == "NZMO2":
                difficulty_bounds = [[1,2],[2,3],[3,4],[4,5],[5,6]]
            elif template == "SMO2":
                difficulty_bounds = [[4,5],[5,6],[6,7],[7,8],[8,9]]
            elif template == "USAMO":
                difficulty_bounds = [[5,7],[7,9],[9,11],[5,7],[7,9],[9,11]]
            elif template == "USAJMO":
                difficulty_bounds = [[3,5],[5,7],[7,8],[3,5],[5,7],[7,8]]
            elif template == "CHINA":
                difficulty_bounds = [[7,8],[8,10],[10,12],[7,8],[8,10],[10,12]]
            elif template == "AFMO": # easter egg
                difficulty_bounds = [[12,"T"],[12,"T"],[12,"T"],[13,"T"]]

        # SMO2 seems to have an unspoken rule to start with geometry at P1 and nowhere else
        if template == "SMO2":
            genre_rule = ["G","ACN","ACN","ACN","ACN"]
        elif template == "IGO":
            genre_rule = ["G","G","G","G","G"]
        else:
            genre_rule = ["ACGN"] * len(difficulty_bounds)

        # pick the genre of each problem
        genres=[]
        while not self.is_genre_legit(genres, template, genre_rule):
            genres=list(map(lambda x: random.choice(x),genre_rule))

        # set up variables
        problems_tex = []
        potds = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=POTD_RANGE).execute().get('values', [])
        already_picked = []

        # render the mock paper
        for i in range(0,len(difficulty_bounds)):
            picked_potd = self.pick_potd(difficulty_bounds[i][0], difficulty_bounds[i][1], genres[i], potds, already_picked, ctx, search_unsolved)
            already_picked.append(picked_potd)
            potd_statement = self.get_potd_statement(int(picked_potd), potds)
            problems_tex.append(f'\\textbf{{Problem {i+1}. (POTD {str(picked_potd)})}}\\\\ ' + potd_statement)
        
        if template in ["IMO","AMO","USAMO","USAJMO","CHINA"] : # 2-day contests
            if template in ["IMO","CHINA","USAMO","USAJMO"]:
                index_day1 = [0,1,2]
                index_day2 = [3,4,5]
            elif template in ["AMO"]:
                index_day1 = [0,1,2,3]
                index_day2 = [4,5,6,7]

            name_day1 = template + ' (Day 1)'
            problems_tex_day1 = [problems_tex[index] for index in index_day1]
            await self.send_out_mock(ctx, name_day1, problems_tex_day1)

            name_day2 = template + ' (Day 2)'
            problems_tex_day2 = [problems_tex[index] for index in index_day2]
            await self.send_out_mock(ctx, name_day2, problems_tex_day2)
        else: # 1-day contests
            await self.send_out_mock(ctx, template, problems_tex)

    @commands.command(aliases=['mock_custom', 'custom_mock'], brief='Create a custom mock paper using past POTDs.',
        help='`-mock_custom [5 7] [7 9] [9 11] [5 7] [7 9] [9 11]`: create a mock paper where problem 1 is d5-7, problem 2 is d7-9, etc.\n'
            '`-mock_custom [3 4 G] [4 5 G] [5 6 G] [6 7 G]`: create a mock paper where problem 1 is d3-4 geometry, problem 2 is d4-5 geometry, etc.')
    @commands.cooldown(1, 30, BucketType.user)
    async def potd_mock_custom(self, ctx, *, rules):
        # parse the user inputed rules
        parsed_rules = self.parse_mock_rules(rules)

        # handle garbage or too long input
        if parsed_rules == False:
            await ctx.send("Custom rule input error! Please input the custom rule like this: `[5 7] [7 9] [9 11]`.")
            return
        if len(parsed_rules) > 15:
            await ctx.send("Maximum number of problems allowed is 15.")
            return

        # get the genre rule
        genre_rule = []
        for parsed_rule in parsed_rules:
            if parsed_rule['genres'] == '':
                genre_rule.append('ACGN')
            else:
                genre_rule.append(parsed_rule['genres'])

        # pick the genre of each problem
        genres=[]
        while not self.is_genre_legit(genres, "Custom", genre_rule):
            genres=list(map(lambda x: random.choice(x),genre_rule))

        # get the difficulty bounds
        difficulty_bounds = []
        for parsed_rule in parsed_rules:
            difficulty_bounds.append([parsed_rule['diff_lower'], parsed_rule['diff_upper']])

        # set up variables
        problems_tex = []
        potds = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=POTD_RANGE).execute().get('values', [])
        already_picked = []
        parsed_rules_string = self.stringify_mock_rules(parsed_rules)

        # render the mock paper
        try:
            for i in range(0,len(difficulty_bounds)):
                picked_potd = self.pick_potd(difficulty_bounds[i][0], difficulty_bounds[i][1], genres[i], potds, already_picked, ctx, True)
                already_picked.append(picked_potd)
                potd_statement = self.get_potd_statement(int(picked_potd), potds)
                problems_tex.append(f'\\textbf{{Problem {i+1}. (POTD {str(picked_potd)})}}\\\\ ' + potd_statement)

            await ctx.send(f'<@{ctx.author.id}> Custom Mock created ({parsed_rules_string})')
            await self.send_out_mock(ctx, '(Custom)', problems_tex)
        except:
            await ctx.send(f'Unable to create mock paper according to custom rule ({parsed_rules_string})')

    async def send_out_mock(self, ctx, name, problems_tex):
        while len(problems_tex) > 0: # still has problems to send out
            title = r'\begin{center}\textbf{\textsf{MODSBot Mock ' + name + r'}}\end{center}'
            problems = ''
            while len(problems_tex) > 0 and len(problems + problems_tex[0]) < 1800 : # add problems one-by-one until no problems left or it's too long
                problems = problems + problems_tex.pop(0) + r'\\ \\'
            problems = problems[0:-5]
            to_tex = f'<@419356082981568522>\n```tex\n {title} {problems}```'
            await ctx.send(to_tex, delete_after=5) 

    def is_genre_legit(self, genres, template, genre_rule):
        if len(genres) != len(genre_rule):
            return False
        
        # the paper should cover as many genre listed in genre_rule as possible
        question_number = len(genre_rule)
        different_genre_number = len(set(''.join(genre_rule)))
        genres_needed = min(question_number, different_genre_number)

        if len(genres) < genres_needed:
            return False

        # the selected genres need to match the genre_rule
        for i in range(0,len(genres)):
            if genres[i] not in genre_rule[i]:
                return False

        if template == "IMO":
            # P3 and P6 should be different genre
            if genres[2] == genres[5]: 
                return False
            
            # The three problems on each day should be different genre
            if len({genres[0],genres[1],genres[2]}) < 3:
                return False
            if len({genres[3],genres[4],genres[5]}) < 3:
                return False

            # Geoff Smith Rule
            genres_geoff_smith = [genres[index] for index in [0,1,3,4]]
            if not ("A" in genres_geoff_smith and "C" in genres_geoff_smith and "G" in genres_geoff_smith and "N" in genres_geoff_smith):
                return False
            
        return True

    def pick_potd(self, diff_lower_bound_filter, diff_upper_bound_filter, genre_filter, potds, already_picked, ctx, search_unsolved:bool):
        solved_potd = []
        if search_unsolved == True:
            get_solved_potd = self.get_potd_solved(ctx)
            get_read_potd = self.get_potd_read(ctx)
            solved_potd = get_solved_potd + get_read_potd

        def match_genre(x,genre_filter):
            for genre in genre_filter:
                if (len(set(x[cfg.Config.config['potd_sheet_genre_col']]).intersection(genre)) == len(genre)):
                    return True
            return False
        
        today = datetime.strptime(datetime.now().strftime("%d %b %Y"), '%d %b %Y')
        
        # filter by genre and difficulty
        if type(diff_upper_bound_filter) == int:
            filtered_potds = [x for x in potds if len(x) > max(cfg.Config.config['potd_sheet_difficulty_col'], cfg.Config.config['potd_sheet_genre_col'])
                            and x[cfg.Config.config['potd_sheet_difficulty_col']].isnumeric()
                            and int(x[cfg.Config.config['potd_sheet_difficulty_col']]) >= diff_lower_bound_filter
                            and int(x[cfg.Config.config['potd_sheet_difficulty_col']]) <= diff_upper_bound_filter                            
                            and match_genre(x,genre_filter)
                            and datetime.strptime(x[cfg.Config.config['potd_sheet_date_col']], '%d %b %Y') < today]
        else: # if diff bound is "T"
            filtered_potds = [x for x in potds if len(x) > max(cfg.Config.config['potd_sheet_difficulty_col'], cfg.Config.config['potd_sheet_genre_col'])
                            and ((x[cfg.Config.config['potd_sheet_difficulty_col']].isnumeric()
                                and int(x[cfg.Config.config['potd_sheet_difficulty_col']]) >= diff_lower_bound_filter)
                                or not x[cfg.Config.config['potd_sheet_difficulty_col']].isnumeric())
                            and match_genre(x,genre_filter)
                            and datetime.strptime(x[cfg.Config.config['potd_sheet_date_col']], '%d %b %Y') < today]


        # pick a POTD
        if len(filtered_potds) > 0:
            filtered_potds_id = list(map(lambda x: int(x[cfg.Config.config['potd_sheet_id_col']]), filtered_potds))
            unsolved_potds_id = [x for x in filtered_potds_id if x not in solved_potd if x not in already_picked]
            if len(unsolved_potds_id) > 0:
                picked_potd = int(random.choice(unsolved_potds_id))
            else:
                not_repeated_potds_id = [x for x in filtered_potds_id if x not in already_picked]
                if len(not_repeated_potds_id) > 0:
                    picked_potd = int(random.choice(not_repeated_potds_id))
                else:
                    picked_potd = int(random.choice(filtered_potds_id))
            return picked_potd
        else:
            return None

    def get_potd_statement(self, number:int, potds):
        current_potd = int(potds[0][0])  # this will be the top left cell which indicates the latest added potd

        if number > current_potd:
            return None

        potd_row = potds[current_potd - number]  # this gets the row requested

        # Create the tex
        potd_statement = ''
        try:
            potd_statement = potd_row[cfg.Config.config['potd_sheet_statement_col']]
            return potd_statement
        except IndexError:
            return None
        
    def parse_mock_rules(self, rules):
        parsed_rules = []

        rules = rules.replace(",", " ")
        res = re.findall(r'\[.*?\]', rules) 

        for substring in res:
            modified_substring = substring[1:-1].split(' ')

            if len(modified_substring) not in [2,3]:
                return False
            if len(modified_substring) == 2:
                modified_substring.append('ACGN')

            try:
                int(modified_substring[0])
                int(modified_substring[1])
            except:
                return False
            if int(modified_substring[0]) > int(modified_substring[1]):
                return False
            
            diff_lower = max(int(modified_substring[0]), 0)
            diff_upper = min(int(modified_substring[1]), 14)
            genres = ''
            possible_genres = ['A', 'C', 'G', 'N']
            for char in modified_substring[2]:
                if char.upper() in possible_genres and char.upper() not in genres:
                    genres += char.upper()
            
            parsed_rule = {
                'diff_lower': diff_lower,
                'diff_upper': diff_upper,
                'genres': genres
            }

            parsed_rules.append(parsed_rule)
        
        return parsed_rules

    def stringify_mock_rules(self, parsed_rules):
        rule_strings = []
        for parse_rule in parsed_rules:
            if parse_rule['genres'] not in ['', 'ACGN']:
                rule_string = f"[{parse_rule['diff_lower']} {parse_rule['diff_upper']} {parse_rule['genres']}]"
            else:
                rule_string = f"[{parse_rule['diff_lower']} {parse_rule['diff_upper']}]"
            rule_strings.append(rule_string)
        return ' '.join(rule_strings)

    @commands.command(aliases=['mark'], brief='Mark the POTD you have solved')
    @commands.cooldown(1, 5, BucketType.user)
    async def potd_mark(self, ctx, *, user_input:str):
        # parse input
        try:
            potd_numbers = [int(i) for i in user_input.split(",")]
        except ValueError:
            await ctx.send("Error: The input contains non-integer values.")
            return

        if len(potd_numbers) > 30:
            await ctx.send("Please don't send more than 30 POTDs in each call.")
            return

        # insert to DB
        added = []
        already_solved = []
        no_potd = []
        no_hint = []
        has_discussion = []
        sheet = self.get_potd_sheet()
        for potd_number in potd_numbers:
            cursor = cfg.db.cursor()
            cursor.execute(f'''SELECT discord_user_id, potd_id, create_date FROM potd_solves 
                                WHERE discord_user_id = {ctx.author.id} 
                                AND potd_id = {potd_number}''')
            result = cursor.fetchall()
            if len(result) > 0:
                already_solved.append(str(potd_number))
            else:
                cursor.execute(f'''INSERT INTO potd_solves (discord_user_id, potd_id, create_date) VALUES
                    ('{ctx.author.id}', '{potd_number}', '{datetime.now()}')''')
                cursor.execute(f'''DELETE FROM potd_read WHERE discord_user_id = {ctx.author.id} AND potd_id = {potd_number}''')
                cursor.execute(f'''DELETE FROM potd_todo WHERE discord_user_id = {ctx.author.id} AND potd_id = {potd_number}''')
                added.append(str(potd_number))
            
            potd_row = self.get_potd_row(potd_number, sheet)
            if potd_row == None or len(potd_row) <= cfg.Config.config['potd_sheet_statement_col']:
                no_potd.append(str(potd_number))
            else:
                if potd_row != None and random.random() <  0.25:
                    if len(potd_row) <= cfg.Config.config['potd_sheet_hint1_col'] or potd_row[cfg.Config.config['potd_sheet_hint1_col']] == None:
                        no_hint.append(str(potd_number))
                if potd_row != None:
                    if len(potd_row) > cfg.Config.config['potd_sheet_discussion_col'] and potd_row[cfg.Config.config['potd_sheet_discussion_col']] != None and potd_row[cfg.Config.config['potd_sheet_discussion_col']] != '':
                        has_discussion.append(str(potd_number))

        # send confirm message
        messages = []
        if len(added) != 0:
            if len(added) == 1:
                messages.append(f'POTD {added[0]} is added to your solved list. Use `-rate {added[0]} <rating>` if you want to rate the difficulty of this problem.')
            else:
                messages.append(f'POTD {",".join(added)} are added to your solved list.')
        if len(already_solved) != 0:
            if len(already_solved) == 1:
                messages.append(f'POTD {already_solved[0]} is already in your solved list.')
            else:
                messages.append(f'POTD {",".join(already_solved)} are already in your solved list.')
        if len(no_potd) != 0:
            if len(no_potd) == 1:
                messages.append(f'There is no POTD {no_potd[0]}. Are you sure you have inputted the correct number?')
            else:
                messages.append(f'There are no POTD  {",".join(no_potd)}. Are you sure you have inputted the correct number?')
        if len(no_hint) != 0:
            if len(no_hint) == 1:
                messages.append(f"There is no hint for POTD {no_hint[0]}. Would you like to contribute one? Contact <@{cfg.Config.config['staffmail_id']}> to submit a hint!")
            else:
                messages.append(f"There are no hint for POTD {','.join(no_hint)}. Would you like to contribute one? Contact <@{cfg.Config.config['staffmail_id']}> to submit a hint!")
        if len(has_discussion) != 0:
            if len(has_discussion) == 1:
                messages.append(f'There is discussion for POTD {has_discussion[0]}. Use `-discussion {has_discussion[0]}` to see the discussion.')
            else:
                messages.append(f"Ther are discussions for POTD {','.join(has_discussion)}. Use `-discussion <number>` to see the discussions.")
        message = "\n".join(messages)
        await ctx.send(message)

    @commands.command(aliases=['unmark'], brief='Unmark the POTD from your solved list')
    @commands.cooldown(1, 5, BucketType.user)
    async def potd_unmark(self, ctx, *, user_input:str):
        # parse input
        try:
            potd_numbers = [int(i) for i in user_input.split(",")]
        except ValueError:
            await ctx.send("Error: The input contains non-integer values.")
            return

        if len(potd_numbers) > 30:
            await ctx.send("Please don't send more than 30 POTDs in each call.")
            return

        # delete from DB
        for potd_number in potd_numbers:
            cursor = cfg.db.cursor()
            cursor.execute(f'''DELETE FROM potd_solves 
                                WHERE discord_user_id = {ctx.author.id} AND potd_id = {potd_number}''')
        
        # send confirm message
        if len(potd_numbers) == 1:
            await ctx.send(f'POTD {potd_numbers[0]} is removed from your solved list. ')
        else:
            await ctx.send(f'POTD {",".join(list(map(str,potd_numbers)))} are removed from your solved list. ')

    @commands.command(aliases=['read'], brief='Mark the POTD you have read')
    @commands.cooldown(1, 5, BucketType.user)
    async def potd_read(self, ctx, *, user_input:str):
        # parse input
        try:
            potd_numbers = [int(i) for i in user_input.split(",")]
        except ValueError:
            await ctx.send("Error: The input contains non-integer values.")
            return

        if len(potd_numbers) > 30:
            await ctx.send("Please don't send more than 30 POTDs in each call.")
            return

        # insert to DB
        added = []
        already_read = []
        no_potd = []
        no_hint = []
        has_discussion = []
        sheet = self.get_potd_sheet()
        for potd_number in potd_numbers:
            cursor = cfg.db.cursor()
            cursor.execute(f'''SELECT discord_user_id, potd_id, create_date FROM potd_read 
                                WHERE discord_user_id = {ctx.author.id} 
                                AND potd_id = {potd_number}''')
            result = cursor.fetchall()
            if len(result) > 0:
                already_read.append(str(potd_number))
            else:
                cursor.execute(f'''INSERT INTO potd_read (discord_user_id, potd_id, create_date) VALUES
                    ('{ctx.author.id}', '{potd_number}', '{datetime.now()}')''')
                cursor.execute(f'''DELETE FROM potd_solves WHERE discord_user_id = {ctx.author.id} AND potd_id = {potd_number}''')
                cursor.execute(f'''DELETE FROM potd_todo WHERE discord_user_id = {ctx.author.id} AND potd_id = {potd_number}''')
                added.append(str(potd_number))
            
            potd_row = self.get_potd_row(potd_number, sheet)
            if potd_row == None or len(potd_row) <= cfg.Config.config['potd_sheet_statement_col']:
                no_potd.append(str(potd_number))
            else:
                if potd_row != None and random.random() <  0.25:
                    if len(potd_row) <= cfg.Config.config['potd_sheet_hint1_col'] or potd_row[cfg.Config.config['potd_sheet_hint1_col']] == None:
                        no_hint.append(str(potd_number))                
                if potd_row != None:
                    if len(potd_row) > cfg.Config.config['potd_sheet_discussion_col'] and potd_row[cfg.Config.config['potd_sheet_discussion_col']] != None and potd_row[cfg.Config.config['potd_sheet_discussion_col']] != '':
                        has_discussion.append(str(potd_number))

        # send confirm message
        messages = []
        if len(added) != 0:
            if len(added) == 1:
                messages.append(f'POTD {added[0]} is added to your read list.')
            else:
                messages.append(f'POTD {",".join(added)} are added to your read list.')
        if len(already_read) != 0:
            if len(already_read) == 1:
                messages.append(f'POTD {already_read[0]} is already in your read list.')
            else:
                messages.append(f'POTD {",".join(already_read)} are already in your read list.')
        if len(no_potd) != 0:
            if len(no_potd) == 1:
                messages.append(f'There is no POTD {no_potd[0]}. Are you sure you have inputted the correct number?')
            else:
                messages.append(f'There are no POTD  {",".join(no_potd)}. Are you sure you have inputted the correct number?')
        if len(no_hint) != 0:
            if len(no_hint) == 1:
                messages.append(f"There is no hint for POTD {no_hint[0]}. Would you like to contribute one? Contact <@{cfg.Config.config['staffmail_id']}> to submit a hint!")
            else:
                messages.append(f"There are no hint for POTD {','.join(no_hint)}. Would you like to contribute one? Contact <@{cfg.Config.config['staffmail_id']}> to submit a hint!")
        if len(has_discussion) != 0:
            if len(has_discussion) == 1:
                messages.append(f'There is discussion for POTD {has_discussion[0]}. Use `-discussion {has_discussion[0]}` to see the discussion.')
            else:
                messages.append(f"Ther are discussions for POTD {','.join(has_discussion)}. Use `-discussion <number>` to see the discussions.")
        message = "\n".join(messages)
        await ctx.send(message)

    @commands.command(aliases=['unread'], brief='Unmark the POTD from your read list')
    @commands.cooldown(1, 5, BucketType.user)
    async def potd_unread(self, ctx, *, user_input:str):
        # parse input
        try:
            potd_numbers = [int(i) for i in user_input.split(",")]
        except ValueError:
            await ctx.send("Error: The input contains non-integer values.")
            return

        if len(potd_numbers) > 30:
            await ctx.send("Please don't send more than 30 POTDs in each call.")
            return

        # delete from DB
        for potd_number in potd_numbers:
            cursor = cfg.db.cursor()
            cursor.execute(f'''DELETE FROM potd_read 
                                WHERE discord_user_id = {ctx.author.id} AND potd_id = {potd_number}''')
        
        # send confirm message
        if len(potd_numbers) == 1:
            await ctx.send(f'POTD {potd_numbers[0]} is removed from your read list. ')
        else:
            await ctx.send(f'POTD {",".join(list(map(str,potd_numbers)))} are removed from your read list. ')

    @commands.command(aliases=['solved'], brief='Show the POTDs you have solved or read',
        help='`-solved`: Show the POTDs you have solved or read.\n'
            '`-solved d`: Show the POTDs you have solved or read, ordered by difficulties.\n'
            '`-solved s`: Show the POTDs you have solved or read, divided into the four subjects.\n')
    @commands.cooldown(1, 5, BucketType.user)
    async def potd_solved(self, ctx, flag=None):
        solved = self.get_potd_solved(ctx)
        read = self.get_potd_read(ctx)
        
        potd_rows = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=POTD_RANGE).execute().get('values', [])
        current_potd = int(potd_rows[0][0])
        
        if len(solved) > 0:
            await self.generate_potd_list_output_string(solved, potd_rows, current_potd, flag, 'solved', ctx)
        if len(read) > 0:
            await self.generate_potd_list_output_string(read, potd_rows, current_potd, flag, 'read', ctx)
        if len(solved) == 0 and len(read) == 0:   
            await ctx.send('Your solved list and read list are empty.')
    
    @commands.command(aliases=['todo'], brief='Mark the POTD into your TODO list')
    @commands.cooldown(1, 5, BucketType.user)
    async def potd_todo(self, ctx, *, user_input:str):
        # parse input
        try:
            potd_numbers = [int(i) for i in user_input.split(",")]
        except ValueError:
            await ctx.send("Error: The input contains non-integer values.")
            return

        if len(potd_numbers) > 30:
            await ctx.send("Please don't send more than 30 POTDs in each call.")
            return

        # insert to DB
        added = []
        already_todo = []
        no_potd = []
        no_hint = []
        has_discussion = []
        sheet = self.get_potd_sheet()
        for potd_number in potd_numbers:
            cursor = cfg.db.cursor()
            cursor.execute(f'''SELECT discord_user_id, potd_id, create_date FROM potd_todo 
                                WHERE discord_user_id = {ctx.author.id} 
                                AND potd_id = {potd_number}''')
            result = cursor.fetchall()
            if len(result) > 0:
                already_todo.append(str(potd_number))
            else:
                cursor.execute(f'''INSERT INTO potd_todo (discord_user_id, potd_id, create_date) VALUES
                    ('{ctx.author.id}', '{potd_number}', '{datetime.now()}')''')
                added.append(str(potd_number))           

        # send confirm message
        messages = []
        if len(added) != 0:
            if len(added) == 1:
                messages.append(f'POTD {added[0]} is added to your TODO list.')
            else:
                messages.append(f'POTD {",".join(added)} are added to your TODO list.')
        if len(already_todo) != 0:
            if len(already_todo) == 1:
                messages.append(f'POTD {already_todo[0]} is already in your TODO list.')
            else:
                messages.append(f'POTD {",".join(already_todo)} are already in your TODO list.')
        message = "\n".join(messages)
        await ctx.send(message)

    @commands.command(aliases=['untodo'], brief='Unmark the POTD from your TODO list')
    @commands.cooldown(1, 5, BucketType.user)
    async def potd_untodo(self, ctx, *, user_input:str):
        # parse input
        try:
            potd_numbers = [int(i) for i in user_input.split(",")]
        except ValueError:
            await ctx.send("Error: The input contains non-integer values.")
            return

        if len(potd_numbers) > 30:
            await ctx.send("Please don't send more than 30 POTDs in each call.")
            return

        # delete from DB
        for potd_number in potd_numbers:
            cursor = cfg.db.cursor()
            cursor.execute(f'''DELETE FROM potd_todo 
                                WHERE discord_user_id = {ctx.author.id} AND potd_id = {potd_number}''')
        
        # send confirm message
        if len(potd_numbers) == 1:
            await ctx.send(f'POTD {potd_numbers[0]} is removed from your TODO list. ')
        else:
            await ctx.send(f'POTD {",".join(list(map(str,potd_numbers)))} are removed from your TODO list. ')

    @commands.command(aliases=['mytodo'], brief='Show the POTDs in your TODO list',
                    help='`-mytodo`: Show the POTDs in your TODO list.\n'
                        '`-mytodo d`: Show the POTDs in your TODO list, ordered by difficulties.\n'
                        '`-mytodo s`: Show the POTDs in your TODO list, divided into the four subjects.\n')
    @commands.cooldown(1, 5, BucketType.user)
    async def potd_mytodo(self, ctx, flag=None):
        todo = self.get_potd_todo(ctx)
        
        potd_rows = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=POTD_RANGE).execute().get('values', [])
        current_potd = int(potd_rows[0][0])
        
        if len(todo) > 0:
            await self.generate_potd_list_output_string(todo, potd_rows, current_potd, flag, 'TODO', ctx, True)
        else:
            await ctx.send('Your TODO list is empty.')

    @commands.command(aliases=['unrated'], brief='Fetch a random POTD that you have solved/read but not yet rated',
                    help='`-unrated`: Fetch a random POTD that you have solved/read but not yet rated.\n')
    @commands.cooldown(1, 5, BucketType.user)
    async def potd_unrated(self, ctx, flag=None):
        solved = self.get_potd_solved(ctx)
        read = self.get_potd_read(ctx)
        rated = self.get_potd_rated(ctx)

        unrated = [x for x in (solved + read) if x not in rated]

        picked_potd = random.choice(unrated)
        await self.potd_fetch(ctx, int(picked_potd))

    @commands.command(aliases=['unrated_list'], brief='Get the list of POTD that you have solved/read but not yet rated',
                    help='`-unrated_list`: Get the list of POTD that you have solved/read but not yet rated.\n'
                        '`-unrated_list d`: Get the list of POTD that you have solved/read but not yet rated, ordered by difficulties.\n'
                        '`-unrated_list s`: Get the list of POTD that you have solved/read but not yet rated, divided into the four subjects.\n')
    @commands.cooldown(1, 5, BucketType.user)
    async def potd_unrated_list(self, ctx, flag=None):
        solved = self.get_potd_solved(ctx)
        read = self.get_potd_read(ctx)
        rated = self.get_potd_rated(ctx)

        solved_unrated = [x for x in solved if x not in rated]
        read_unrated = [x for x in read if x not in rated]

        potd_rows = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=POTD_RANGE).execute().get('values', [])
        current_potd = int(potd_rows[0][0])
        
        if len(solved_unrated) > 0:
            await self.generate_potd_list_output_string(solved_unrated, potd_rows, current_potd, flag, 'unrated (solved)', ctx, True)
        if len(read_unrated) > 0:
            await self.generate_potd_list_output_string(read_unrated, potd_rows, current_potd, flag, 'unrated (read)', ctx, True)
        if len(solved_unrated) == 0 and len(read_unrated) == 0:
            await ctx.send('You have no unrated POTD.')

    async def generate_potd_list_output_string(self, potd_list, potd_rows, current_potd, flag, adjective, ctx, show_total=True):
        if flag == "d":
            solved_by_difficulty = {}
            for number in potd_list:
                if number > current_potd or number <= 0:
                    difficulty = "(Unknown)"
                else:
                    potd_row = potd_rows[current_potd - number]
                    if len(potd_row) > cfg.Config.config['potd_sheet_difficulty_col']:
                        difficulty = potd_row[cfg.Config.config['potd_sheet_difficulty_col']]
                    else:
                        difficulty = "(Unknown)"

                if difficulty not in solved_by_difficulty:
                    solved_by_difficulty[difficulty] = []
                solved_by_difficulty[difficulty].append(number)            
            
            sorted_keys = sorted(solved_by_difficulty.keys(), key=lambda x: (x.isnumeric(),int(x) if x.isnumeric() else x), reverse=True)
            solved_by_difficulty = {key:solved_by_difficulty[key] for key in sorted_keys}

            output_string = f'__**Your {adjective} POTD**__ \n'
            for key in solved_by_difficulty:
                if show_total == True:
                    total = len([potd for potd in potd_rows if len(potd) > cfg.Config.config['potd_sheet_difficulty_col']
                                and potd[cfg.Config.config['potd_sheet_difficulty_col']] == key])
                    output_string += "**D" + key + ":** " + f"{solved_by_difficulty[key]} ({len(solved_by_difficulty[key])}/{total})" + "\n"
                else:
                    output_string += "**D" + key + ":** " + f"{solved_by_difficulty[key]} " + "\n"
            if show_total == True:
                output_string += f"(Total: {len(potd_list)}/{len(potd_rows)})"
            await self.send_potd_solved(ctx, output_string)
        elif flag == "s":
            solved_by_genre = {'A':[], 'C':[], 'G':[], 'N':[]}
            for number in potd_list:
                if number > current_potd or number <= 0:
                    genre = "(Unknown)"
                else:
                    potd_row = potd_rows[current_potd - number]
                    if len(potd_row) > cfg.Config.config['potd_sheet_genre_col']:
                        genre = potd_row[cfg.Config.config['potd_sheet_genre_col']]
                    else:
                        genre = "(Unknown)"

                if 'A' in genre:
                    solved_by_genre['A'].append(number)
                if 'C' in genre:
                    solved_by_genre['C'].append(number)
                if 'G' in genre:
                    solved_by_genre['G'].append(number)
                if 'N' in genre:
                    solved_by_genre['N'].append(number)

            output_string = f'__**Your {adjective} POTD**__ \n'
            for key in solved_by_genre:
                if show_total == True:
                    total = len([potd for potd in potd_rows if len(potd) > cfg.Config.config['potd_sheet_difficulty_col']
                                and key in potd[cfg.Config.config['potd_sheet_genre_col']]])
                    output_string += "**" + key + ":** " + f"{solved_by_genre[key]} ({len(solved_by_genre[key])}/{total})" + "\n"
                else:
                    output_string += "**" + key + ":** " + f"{solved_by_genre[key]} " + "\n"
            if show_total == True:
                output_string += f"(Total: {len(potd_list)}/{len(potd_rows)})"
            await self.send_potd_solved(ctx, output_string)
        else:
            if show_total == True:
                output_string = f'__**Your {adjective} POTD**__ \n{potd_list}' + "\n"
            else:
                output_string = f'__**Your {adjective} POTD**__ \n{potd_list}' + "\n"
            if show_total == True:
                output_string += f"(Total: {len(potd_list)}/{len(potd_rows)})"
            await self.send_potd_solved(ctx, output_string)
        
    
    def get_potd_solved(self, ctx):
        cursor = cfg.db.cursor()
        cursor.execute(f'''SELECT discord_user_id, potd_id, create_date FROM potd_solves 
                            WHERE discord_user_id = {ctx.author.id} 
                            ORDER BY potd_id DESC''')
        return [x[1] for x in cursor.fetchall()]
    
    def get_potd_read(self, ctx):
        cursor = cfg.db.cursor()
        cursor.execute(f'''SELECT discord_user_id, potd_id, create_date FROM potd_read 
                            WHERE discord_user_id = {ctx.author.id} 
                            ORDER BY potd_id DESC''')
        return [x[1] for x in cursor.fetchall()]
    
    def get_potd_todo(self, ctx):
        cursor = cfg.db.cursor()
        cursor.execute(f'''SELECT discord_user_id, potd_id, create_date FROM potd_todo
                            WHERE discord_user_id = {ctx.author.id} 
                            ORDER BY potd_id DESC''')
        return [x[1] for x in cursor.fetchall()]
    
    def get_potd_rated(self, ctx):
        cursor = cfg.db.cursor()
        cursor.execute(f'SELECT * FROM ratings WHERE userid = {ctx.author.id}')
        return [x[1] for x in cursor.fetchall()]

    # send message in batches of 1900+e characters because of 2k character limit
    async def send_potd_solved(self, ctx, output_string):
        i = 0
        output_batch = ""
        while i < len(output_string):
            if output_batch == "":
                jump = min(1900, len(output_string)-i)
                output_batch += output_string[i:i+jump]
                i += jump
            else:
                output_batch += output_string[i]
                i += 1
            if output_batch[-1] == "," or output_batch[-1] == "]" or len(output_batch) == 2000 or i==len(output_string): # we end a batch at "," or "]"
                await ctx.send(output_batch)
                output_batch = ""


    @commands.command(aliases=['hint'], brief='Get hint for the POTD.')
    @commands.cooldown(1, 10, BucketType.user)
    async def potd_hint(self, ctx, number: int, hint_number: int = 1):
        sheet = self.get_potd_sheet()
        potd_row = self.get_potd_row(number, sheet)
        if potd_row == None:
            await ctx.send(f"There is no potd for day {number}. ")
            return
        else:  
            if hint_number == 1:
                if len(potd_row) <= cfg.Config.config['potd_sheet_hint1_col'] or potd_row[cfg.Config.config['potd_sheet_hint1_col']] == None or potd_row[cfg.Config.config['potd_sheet_hint1_col']] == '':
                    await ctx.send(f"There is no hint for POTD {number}. Would you like to contribute one? Contact <@{cfg.Config.config['staffmail_id']}> to submit a hint!")
                    return
                else:
                    await ctx.send(f"Hint for POTD {number}:\n")
                    await ctx.send(f"<@{cfg.Config.config['paradox_id']}> texsp \n||```latex\n{potd_row[cfg.Config.config['potd_sheet_hint1_col']]}```||")
                    if len(potd_row) > cfg.Config.config['potd_sheet_hint2_col'] and potd_row[cfg.Config.config['potd_sheet_hint2_col']] != None and potd_row[cfg.Config.config['potd_sheet_hint2_col']] != '':
                        await ctx.send(f"There is another hint for this POTD. Use `-hint {number} 2` to get the hint.")
            elif hint_number == 2:
                if len(potd_row) <= cfg.Config.config['potd_sheet_hint2_col'] or potd_row[cfg.Config.config['potd_sheet_hint2_col']] == None or potd_row[cfg.Config.config['potd_sheet_hint2_col']] == '':
                    await ctx.send(f"There is no hint 2 for POTD {number}. Would you like to contribute one? Contact <@{cfg.Config.config['staffmail_id']}> to submit a hint!")
                    return
                else:
                    await ctx.send(f"Hint 2 for POTD {number}:\n")
                    await ctx.send(f"<@{cfg.Config.config['paradox_id']}> texsp \n||```latex\n{potd_row[cfg.Config.config['potd_sheet_hint2_col']]}```||")
                    if len(potd_row) > cfg.Config.config['potd_sheet_hint3_col'] and potd_row[cfg.Config.config['potd_sheet_hint3_col']] != None and potd_row[cfg.Config.config['potd_sheet_hint3_col']] != '':
                        await ctx.send(f"There is another hint for this POTD. Use `-hint {number} 3` to get the hint.")
            elif hint_number == 3:
                if len(potd_row) <= cfg.Config.config['potd_sheet_hint3_col'] or potd_row[cfg.Config.config['potd_sheet_hint3_col']] == None or potd_row[cfg.Config.config['potd_sheet_hint3_col']] == '':
                    await ctx.send(f"There is no hint 3 for POTD {number}. Would you like to contribute one? Contact <@{cfg.Config.config['staffmail_id']}> to submit a hint!")
                    return
                else:
                    await ctx.send(f"Hint 3 for POTD {number}:\n")
                    await ctx.send(f"<@{cfg.Config.config['paradox_id']}> texsp \n||```latex\n{potd_row[cfg.Config.config['potd_sheet_hint3_col']]}```||")
            else:
                await ctx.send("Hint number should be from 1 to 3.")

    @commands.command(aliases=['answer'], brief='Get answer for the POTD.')
    @commands.cooldown(1, 10, BucketType.user)
    async def potd_answer(self, ctx, number: int):
        sheet = self.get_potd_sheet()
        potd_row = self.get_potd_row(number, sheet)
        if potd_row == None:
            await ctx.send(f"There is no potd for day {number}. ")
            return
        else:
            if len(potd_row) <= cfg.Config.config['potd_sheet_answer_col'] or potd_row[cfg.Config.config['potd_sheet_answer_col']] == None or potd_row[cfg.Config.config['potd_sheet_answer_col']] == '':
                await ctx.send(f"There is no answer provided for POTD {number}. Would you like to contribute one? Contact <@{cfg.Config.config['staffmail_id']}> to submit your answer!")
                return
            else:
                await ctx.send(f"Answer for POTD {number}:\n")
                await ctx.send(f"<@{cfg.Config.config['paradox_id']}> texsp \n||```latex\n{potd_row[cfg.Config.config['potd_sheet_answer_col']]}```||")

    @commands.command(aliases=['discussion'], brief='Get discussion for the POTD.')
    @commands.cooldown(1, 10, BucketType.user)
    async def potd_discussion(self, ctx, number: int):
        sheet = self.get_potd_sheet()
        potd_row = self.get_potd_row(number, sheet)
        if potd_row == None:
            await ctx.send(f"There is no potd for day {number}. ")
            return
        else:
            if len(potd_row) <= cfg.Config.config['potd_sheet_discussion_col'] or potd_row[cfg.Config.config['potd_sheet_discussion_col']] == None or potd_row[cfg.Config.config['potd_sheet_discussion_col']] == '':
                await ctx.send(f"There is no discussion provided for POTD {number}.")
                return
            else:
                await ctx.send(f"Discussion for POTD {number}:\n")
                await ctx.send(f"<@{cfg.Config.config['paradox_id']}> texsp \n||```latex\n{potd_row[cfg.Config.config['potd_sheet_discussion_col']]}```||")

    @commands.command(aliases=['solution'], brief='Get solution for the POTD.')
    @commands.cooldown(1, 10, BucketType.user)
    async def potd_solution(self, ctx, number: int):
        sheet = self.get_potd_sheet()
        potd_row = self.get_potd_row(number, sheet)
        if potd_row == None:
            await ctx.send(f"There is no potd for day {number}. ")
            return
        else:
            if len(potd_row) <= cfg.Config.config['potd_sheet_solution_col'] or potd_row[cfg.Config.config['potd_sheet_solution_col']] == None or potd_row[cfg.Config.config['potd_sheet_solution_col']] == '':
                solution = None
            else:
                solution = potd_row[cfg.Config.config['potd_sheet_solution_col']]
            if len(potd_row) <= cfg.Config.config['potd_sheet_solution_link_col'] or potd_row[cfg.Config.config['potd_sheet_solution_link_col']] == None or potd_row[cfg.Config.config['potd_sheet_solution_link_col']] == '':
                solution_link = None
            else:
                solution_link = potd_row[cfg.Config.config['potd_sheet_solution_link_col']]

            if solution == None and solution_link == None:
                await ctx.send(f"There is no solution provided for POTD {number}. Would you like to contribute one? Contact <@{cfg.Config.config['staffmail_id']}> to submit your solution!")
                return
            else:
                if solution != None:
                    await ctx.send(f"Solution for POTD {number}:\n")
                    await ctx.send(f"<@{cfg.Config.config['paradox_id']}> texsp \n||```latex\n{potd_row[cfg.Config.config['potd_sheet_solution_col']]}```||")
                if solution_link != None:
                    await ctx.send(f"Solution Link for POTD {number}:\n{potd_row[cfg.Config.config['potd_sheet_solution_link_col']]}")

    def get_potd_sheet(self):
        sheet = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_sheet'],
                                                               range=POTD_RANGE).execute()
        return sheet

    def get_potd_row(self, number, sheet):
        values = sheet.get('values', [])
        current_potd = int(values[0][0])  # this will be the top left cell which indicates the latest added potd

        if number > current_potd or number < 1:
            return None

        try:
            potd_row = values[current_potd - number]  # this gets the row requested
            return potd_row
        except IndexError:
            return None

    @commands.command(aliases=['remove_potd'], brief='Deletes the potd with the provided number. ')
    @commands.check(is_pc)
    async def delete_potd(self, ctx, number: int):
        
        # It can only handle one at a time!
        if not self.listening_in_channel in [-1, -2]:
            await dm_or_channel(ctx.author, self.bot.get_channel(cfg.Config.config['helper_lounge']),
                "Please wait until the previous call has finished!")
            return
        self.listening_in_channel = 0
        
        # Delete old POTD
        cursor = cfg.db.cursor()
        cursor.execute(f"SELECT problem_msg_id, source_msg_id, ping_msg_id FROM potd_info WHERE potd_id = '{number}'")
        result = cursor.fetchall()
        cursor.execute(f"DELETE FROM potd_info WHERE potd_id = '{number}'")
        cfg.db.commit()
        for i in result:
            for j in i:
                try:
                    await self.bot.get_channel(cfg.Config.config['potd_channel']).get_partial_message(int(j)).delete()
                except Exception:
                    pass
        self.listening_in_channel = -1

    @commands.command(aliases=['update_potd'], brief='Replaces the potd with the provided number. ')
    @commands.check(is_pc)
    async def replace_potd(self, ctx, number: int):
        
        # It can only handle one at a time!
        if not self.listening_in_channel == -1:
            await dm_or_channel(ctx.author, self.bot.get_channel(cfg.Config.config['helper_lounge']),
                "Please wait until the previous call has finished!")
            return

        await self.delete_potd(ctx, number)
        await self.potd_display(ctx, number)

    @commands.command(aliases=['rate'], brief='Rates a potd based on difficulty. ')
    async def potd_rate(self, ctx, potd: int, rating: int, overwrite: bool = False):
        if rating < 0 or rating > 14:
            await ctx.send(f'<@{ctx.author.id}> POTD rating is only allowed from 0 to 14.')
            return

        # Delete messages if it's in a guild
        if ctx.guild is not None:
            await ctx.message.delete()

        cursor = cfg.db.cursor()
        cursor.execute(f'SELECT * FROM ratings where prob = {potd} and userid = {ctx.author.id} LIMIT 1')
        result = cursor.fetchone()
        # print(result)
        if result is None:
            sql = 'INSERT INTO ratings (prob, userid, rating) VALUES (?, ?, ?)'
            cursor.execute(sql, (potd, ctx.author.id, rating))
            cfg.db.commit()
            await ctx.send(f'<@{ctx.author.id}> You have rated POTD {potd} d||{rating}  ||.')
        else:
            if not overwrite:
                await ctx.send(
                    f'<@{ctx.author.id}> You already rated this POTD d||{result[3]}  ||. '
                    f'If you wish to overwrite append `True` to your previous message, like `-rate {potd} <rating> True` ')
            else:
                cursor.execute(f'UPDATE ratings SET rating = {rating} WHERE idratings = {result[0]}')
                cfg.db.commit()
                await ctx.send(f'<@{ctx.author.id}> You have rated POTD {potd} d||{rating}  ||.')
        await self.edit_source(potd)

    @commands.command(aliases=['rating'], brief='Finds the median of a POTD\'s ratings')
    async def potd_rating(self, ctx, potd: int, full: bool = True):
        cursor = cfg.db.cursor()
        cursor.execute(f'SELECT * FROM ratings WHERE prob = {potd} ORDER BY rating')
        result = cursor.fetchall()
        if len(result) == 0:
            await ctx.send(f'No ratings for POTD {potd} yet. ')
        else:
            median = statistics.median([row[3] for row in result])
            await ctx.send(f'Median community rating for POTD {potd} is d||{median}  ||. ')
            if full:
                embed = discord.Embed()
                embed.add_field(name=f'Full list of community rating for POTD {potd}',
                    value='\n'.join([f'<@!{row[2]}>: d||{row[3]}  ||' for row in result]))
                await ctx.send(embed=embed)

    @commands.command(aliases=['myrating'], brief='Checks your rating of a potd. ')
    async def potd_rating_self(self, ctx, potd: int):
        cursor = cfg.db.cursor()
        cursor.execute(f'SELECT * FROM ratings WHERE prob = {potd} AND userid = {ctx.author.id}')
        result = cursor.fetchone()
        if result is None:
            await ctx.author.send(f'You have not rated potd {potd}. ')
        else:
            await ctx.author.send(f'You have rated potd {potd} as difficulty level {result[3]}')

    @commands.command(aliases=['myratings'], brief='Checks all your ratings. ')
    async def potd_rating_all(self, ctx):
        cursor = cfg.db.cursor()
        cursor.execute(f'SELECT * FROM ratings WHERE userid = {ctx.author.id}')
        result = cursor.fetchall()
        if len(result) == 0:
            await ctx.author.send('You have not rated any problems!')
        else:
            ratings = '\n'.join([f'{i[1]:<6}{i[3]}' for i in result])
            await ctx.author.send(f'Your ratings: ```Potd  Rating\n{ratings}```You have rated {len(result)} potds. ')

    @commands.command(aliases=['rmrating', 'unrate'], brief='Removes your rating for a potd. ')
    async def potd_rating_remove(self, ctx, potd: int):
        cursor = cfg.db.cursor()
        cursor.execute(f'SELECT * FROM ratings WHERE prob = {potd} AND userid = {ctx.author.id}')
        result = cursor.fetchone()
        if result is None:
            await ctx.author.send(f'You have not rated potd {potd}. ')
        else:
            cursor.execute(f'DELETE FROM ratings WHERE prob = {potd} AND userid = {ctx.author.id}')
            await ctx.author.send(f'Removed your rating of difficulty level {result[3]} for potd {potd}. ')
            await self.edit_source(potd)

    def potd_notif_embed(self, ctx, colour):

        result = None
        def subcriteria(a):
            if result[1][a] == 'x':
                return 'Off'
            else:
                return f'D{int(result[1][a:a+2])}-{int(result[1][a+2:a+4])}'

        cursor = cfg.db.cursor()
        cursor.execute(f'SELECT * FROM potd_ping2 WHERE user_id = {ctx.author.id}')
        result = cursor.fetchone()
        if result is None:
            return None
        embed = discord.Embed(colour=colour)
        try:
            if ctx.author.nick is None:
                embed.add_field(name='Username', value=ctx.author.name)
            else:
                embed.add_field(name='Nickname', value=ctx.author.nick)
        except Exception:
            embed.add_field(name='Username', value=ctx.author.name)
        for i in range(4):
            embed.add_field(name=['Algebra', 'Combinatorics', 'Geometry', 'Number Theory'][i], value=subcriteria(4*i))
        embed.set_footer(text='Use `-help pn` for help. ')
        return embed

    @commands.command(aliases=['pn'], brief='Customizes potd pings. ', help='`-pn`: enable POTD notifications or show settings\n'
                                                                            '`-pn a1-7`: set difficulty range for category\n'
                                                                            '`-pn c`: toggle notifications for category\n'
                                                                            '`-pn a1-7 c`: combine commands\n'
                                                                            '`-pn off`: disable notifications')
    async def potd_notif(self, ctx, *criteria:str):

        # Empty criteria
        cursor = cfg.db.cursor()
        criteria = list(criteria)
        if len(criteria) == 0:
            cursor.execute(f"SELECT * FROM potd_ping2 WHERE user_id = '{ctx.author.id}'")
            result = cursor.fetchone()
            if result == None:
                cursor.execute(f'''INSERT INTO potd_ping2 (user_id, criteria)
                    VALUES('{ctx.author.id}', '0 120 120 120 12')''')
                cfg.db.commit()
                await ctx.send('Your POTD notification settings have been updated: ', embed=self.potd_notif_embed(ctx, 0x5FE36A))
            else:
                await ctx.send('Here are your POTD notification settings: ', embed=self.potd_notif_embed(ctx, 0xDCDCDC))
            return

        # Turn off ping
        if criteria[0].lower() == 'off':
            cursor.execute(f"DELETE FROM potd_ping2 WHERE user_id = '{ctx.author.id}'")
            cfg.db.commit()
            await ctx.send('Your POTD notifications have been turned off. ')
            return

        # Run criteria
        cursor.execute(f"SELECT * FROM potd_ping2 WHERE user_id = '{ctx.author.id}'")
        result = cursor.fetchone()
        if result == None:
            cursor.execute(f'''INSERT INTO potd_ping2 (user_id, criteria)
                VALUES('{ctx.author.id}', 'xxxxxxxxxxxxxxxx')''')
            cursor.execute(f"SELECT * FROM potd_ping2 WHERE user_id = '{ctx.author.id}'")
            result = cursor.fetchone()
        result = list(result)

        temp = "".join(criteria).lower()
        criteria = [temp[0]]
        for i in temp[1:]:
            if i in ['a', 'c', 'g', 'n']:
                criteria.append(i)
            else:
                criteria[len(criteria)-1] += i
        
        # Difficulty only
        if len(criteria) == 1:
            temp = criteria[0].split('-')
            if len(temp) == 2:
                try:
                    min = int(temp[0])
                    max = int(temp[1])
                    if (0 <= min <= max <= 12):
                        if result[1] == 'xxxxxxxxxxxxxxxx':
                            result[1] = '                '
                        temp = ''
                        for i in range(4):
                            if result[1][4*i] == 'x':
                                temp += 'xxxx'
                            else:
                                temp += str(min).ljust(2) + str(max).ljust(2)
                        cursor.execute(f"UPDATE potd_ping2 SET criteria = '{temp}' WHERE user_id = '{ctx.author.id}'")
                        cfg.db.commit()
                        await ctx.send('Your POTD notification settings have been updated: ', embed=self.potd_notif_embed(ctx, 0x5FE36A))
                        return
                    else:
                        cfg.db.rollback()
                        await ctx.send(f'`{criteria[0]}` Invalid difficulty range! ')
                        return
                except ValueError:
                    pass

        remaining = ['a', 'c', 'g', 'n']
        for i in criteria:
            if i in remaining:
                # Category without difficulty
                remaining.remove(i)
                index = ['a', 'c', 'g', 'n'].index(i[0])
                if result[1][4*index] == 'x':
                    result[1] = result[1][:4*index] + '0 12' + result[1][4*index+4:]
                else:
                    result[1] = result[1][:4*index] + 'xxxx' + result[1][4*index+4:]
            else:
                # Category with difficulty
                criterion = i[1:].split('-')
                if (i[0] not in remaining) or (len(criterion) != 2):
                    cfg.db.rollback()
                    await ctx.send(f'`{i}` Invalid input format! ')
                    return
                try:
                    min = int(criterion[0])
                    max = int(criterion[1])
                    if not (0 <= min <= max <= 12):
                        cfg.db.rollback()
                        await ctx.send(f'`{i}` Invalid difficulty range! ')
                        return
                except ValueError:
                    cfg.db.rollback()
                    await ctx.send(f'`{i}` Invalid input format! ')
                    return
                remaining.remove(i[0])
                index = ['a', 'c', 'g', 'n'].index(i[0])
                result[1] = f'{result[1][:4*index]}{str(min).ljust(2)}{str(max).ljust(2)}{result[1][4*index+4:]}'

        cursor.execute(f"UPDATE potd_ping2 SET criteria = '{result[1]}' WHERE user_id = '{ctx.author.id}'")
        cfg.db.commit()
        await ctx.send('Your POTD notification settings have been updated: ', embed=self.potd_notif_embed(ctx, 0x5FE36A))

    @commands.command()
    @commands.check(cfg.is_staff)
    async def enable_potd_dm(self, ctx, status:bool=None):
        if status is None:
            self.enable_dm = not self.enable_dm
        else:
            self.enable_dm = status
        cursor = cfg.db.cursor()
        cursor.execute(f"UPDATE settings SET value = '{str(self.enable_dm)}' WHERE setting = 'potd_dm'")
        cfg.db.commit()
        await self.bot.get_channel(cfg.Config.config['log_channel']).send(
            f'**POTD notifications set to `{self.enable_dm}` by {ctx.author.nick} ({ctx.author.id})**')

    def post_proposed_potd(self):
        self.bot.loop.create_task(self.post_proposed_potd_task())
    
    async def post_proposed_potd_task(self):
        # Read from spreadsheet
        proposed_problems = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_proposal_sheet'],
                                                                           range='A:M').execute().get('values', [])

        for i, problem in enumerate(proposed_problems):
            # Find unposted problems
            if len(problem) < 12 or problem[11] == "":
                number = i
                user = problem[1]
                user_id = problem[2]
                problem_statement = problem[3]
                source = problem[4]
                genre = problem[5]
                difficulty = problem[6]
                hint1 = problem[7]
                try:
                    hint2 = problem[8]
                except:
                    hint2 = ""
                try:
                    hint3 = problem[9]
                except:
                    hint3 = ""
                try:
                    proposer_msg = problem[10]
                except:
                    proposer_msg = ""

                # Post in forum
                forum = self.bot.get_channel(cfg.Config.config['potd_proposal_forum'])
                post_result = await forum.create_thread(name=f"POTD Proposal #{number} from {user}", 
                                          content=f"POTD Proposal #{number} from {user} <@!{user_id}> ({user_id})\nProblem Statement: ```latex\n{problem_statement}\n```",
                                          applied_tags=[forum.get_tag(cfg.Config.config['potd_proposal_forum_tag_pending'])]) 
                thread = post_result[0]
                
                problem_info = f"Source: ||{source}|| \n" + f"Genre: ||{genre}  || \n" + f"Difficulty: ||{difficulty}  ||"
                if proposer_msg != "" and proposer_msg != None:
                    problem_info += f"\nProposer's message: {proposer_msg}\n"
                await thread.send(problem_info)
                await asyncio.sleep(10)

                await thread.send(f"Hint 1:")
                await thread.send(f"<@{cfg.Config.config['paradox_id']}> texsp\n||```latex\n{hint1}```||")
                await asyncio.sleep(10)
                if hint2 != "" and hint2 != None:
                    await thread.send(f"Hint 2:")
                    await thread.send(f"<@{cfg.Config.config['paradox_id']}> texsp\n||```latex\n{hint2}```||")
                    await asyncio.sleep(10)
                if hint3 != "" and hint3 != None:
                    await thread.send(f"Hint 3:")
                    await thread.send(f"<@{cfg.Config.config['paradox_id']}> texsp\n||```latex\n{hint3}```||")
                    await asyncio.sleep(10)

                # Mark problem as posted
                request = cfg.Config.service.spreadsheets().values().update(spreadsheetId=cfg.Config.config['potd_proposal_sheet'], 
                                                                            range=f'L{i+1}', valueInputOption='RAW',body={"range": f'L{i+1}', "values": [["Y"]] })
                response = request.execute()

                # Mark thread ID
                request = cfg.Config.service.spreadsheets().values().update(spreadsheetId=cfg.Config.config['potd_proposal_sheet'], 
                                                                            range=f'M{i+1}', valueInputOption='RAW',body={"range": f'M{i+1}', "values": [[str(thread.id)]] })
                response = request.execute()

                # Send notification to proposer
                try:
                    guild = self.bot.get_guild(cfg.Config.config['mods_guild'])
                    member = guild.get_member(int(user_id))
                    if member is not None and not member.bot:
                        await member.send(f"Hi! We have received your POTD Proposal `{source}`. Thanks for your submission!")
                except Exception as e:
                    print(e)

    @commands.command()
    @commands.check(is_pc)
    async def potd_pending(self, ctx, number: int):
        await self.potd_proposal_status_change(ctx, number, "Pending")
        await ctx.send(f"POTD Proposal #{number} status modified to Pending")

    @commands.command()
    @commands.check(is_pc)
    async def potd_accept(self, ctx, number: int):
        await self.potd_proposal_status_change(ctx, number, "Accepted")
        await ctx.send(f"POTD Proposal #{number} status modified to Accepted")
    
    @commands.command()
    @commands.check(is_pc)
    async def potd_reject(self, ctx, number: int):
        await self.potd_proposal_status_change(ctx, number, "Rejected")
        await ctx.send(f"POTD Proposal #{number} status modified to Rejected")

    async def potd_proposal_status_change(self, ctx, number: int, status):
        tag_id = 0
        if status == "Pending":
            tag_id = cfg.Config.config['potd_proposal_forum_tag_pending']
        elif status == "Accepted":
            tag_id = cfg.Config.config['potd_proposal_forum_tag_accepted']
        elif status == "Rejected":
            tag_id = cfg.Config.config['potd_proposal_forum_tag_rejected']

        # Load the proposal sheet
        proposed_problems = cfg.Config.service.spreadsheets().values().get(spreadsheetId=cfg.Config.config['potd_proposal_sheet'],
                                                                           range='A:M').execute().get('values', [])

        # Edit the thread tag
        forum = self.bot.get_channel(cfg.Config.config['potd_proposal_forum'])
        row = number
        thread_id = proposed_problems[row][12]
        thread = ctx.guild.get_thread(int(thread_id))
        await thread.edit(applied_tags=[forum.get_tag(tag_id)])        

    # manually invoke the proposal check
    @commands.command()
    @commands.check(cfg.is_mod_or_tech)
    async def potd_proposal(self, ctx):
        self.bot.loop.create_task(self.post_proposed_potd_task())


    # scan potd channel for image link
    @commands.command()
    @commands.check(cfg.is_mod_or_tech)
    async def potd_image_scan(self, ctx, begin: int, end: int, write: bool = False):
        sheet = self.get_potd_sheet()
        potd_channel = self.bot.get_channel(cfg.Config.config['potd_channel'])
        image_links = []
        for number in range(begin, end+1):
            try:
                potd_row = self.get_potd_row(number, sheet)
                date_start = datetime.strptime(potd_row[cfg.Config.config['potd_sheet_date_col']], '%d %b %Y')
                date_end = date_start + timedelta(hours=23)
                messages = [message async for message in potd_channel.history(limit=10, after=date_start, before=date_end, oldest_first=True)]
                paradox_messages = [x for x in messages if x.author.id==cfg.Config.config['paradox_id']]
                potd_message = paradox_messages[0]
                
                image_link = str(potd_message.attachments[0].proxy_url)
                image_links.append(image_link)

                if write:
                    # record the link to rendered image if it is in POTD channel 
                    # get the row and column to update
                    column = openpyxl.utils.get_column_letter(cfg.Config.config['potd_sheet_image_link_col']+1)
                    values = sheet.get('values', [])
                    current_potd = int(values[0][0])  # this will be the top left cell which indicates the latest added potd
                    row = current_potd - number + 2  # this gets the row requested
                    # update the source_msg in the sheet
                    request = cfg.Config.service.spreadsheets().values().update(spreadsheetId=cfg.Config.config['potd_sheet'], 
                                                                range=f'{column}{row}', valueInputOption='RAW',body={"range": f'{column}{row}', "values": [[image_link]] })
                    response = request.execute()
            except:
                pass

        for image_link in image_links:
            async with aiohttp.ClientSession() as session:
                async with session.get(image_link) as resp:
                    if resp.status != 200:
                        return await ctx.send('Could not download file...')
                    data = io.BytesIO(await resp.read())
                    await ctx.send(file=discord.File(data, f'potd.png'))


async def setup(bot):
    await bot.add_cog(Potd(bot))
